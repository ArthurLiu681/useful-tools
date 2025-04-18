import pandas
import os
from multiprocessing import Process, Lock, Manager
import openpyxl
import datetime

def read_excel_chunk(file_path, sheetname, minr, maxr, cc, datas, lock, key):
    print("%d begin read: " % cc + datetime.datetime.now().strftime("%H:%M:%S") + "\n")
    sheet = openpyxl.load_workbook(file_path, read_only=True, data_only=True)[sheetname]
    data_list = []
    for row in sheet.iter_rows(min_row=minr, max_row=maxr, values_only=True):
        data_list.append(list(row))
    data = pandas.DataFrame(data_list)
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        data.columns = row
    data.fillna("", inplace=True)
    data = data[data[key] != ""].reset_index(drop=True)
    if None in list(data.columns):
        data = data.drop(columns=[None], axis=1)
    print("%d finish read: " % cc + datetime.datetime.now().strftime("%H:%M:%S")+"\n")
    lock.acquire()
    datas[cc] = data
    lock.release()
    print("%d finished add to list: " % cc + datetime.datetime.now().strftime("%H:%M:%S")+"\n")

def write_df(desDir, lt):
    print("%s to %s begin write: " % (lt[0][0],lt[-1][0]) + datetime.datetime.now().strftime("%H:%M:%S")+"\n")
    for company, data in lt:
        data.to_excel(os.path.join(desDir, company.replace("?", " ") + "-2024收入成本大表.xlsx"), index=False)
    print("%s to %s finish write: " % (lt[0][0], lt[-1][0]) + datetime.datetime.now().strftime("%H:%M:%S")+"\n")

if __name__ == '__main__':
    key = "公司"
    read_nt = 5
    write_nt = 10
    manager = Manager()
    datas = manager.dict()
    c = 0
    lock = Lock()
    processes = []
    directory = 'C:\\Users\\0209105\\Desktop\\path'
    desdir = 'C:\\Users\\0209105\\Desktop\\despath'
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        file = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheetname in file.sheetnames:
            sheet = file[sheetname]
            tc = (sheet.max_row-1) // read_nt
            for j in range(read_nt):
                processes.append(Process(target=read_excel_chunk, args=(file_path, sheetname, 1+j*tc+1, 1+j*tc+tc, c,datas,lock,key,)))
                c = c+1
                processes[-1].start()
            if sheet.max_row-1-read_nt*tc > 0:
                processes.append(Process(target=read_excel_chunk, args=(file_path, sheetname, 1+read_nt*tc+1, sheet.max_row, c,datas,lock,key,)))
                c = c + 1
                processes[-1].start()
    
    for process in processes:
        process.join()
    
    df = datas[0]
    for i in range(1, c):
        df = pandas.concat([df, datas[i]])
        if datas[i].shape[0] > 0:
            print("%s to %s concated\n" %(list(datas[i]["序号"])[0], list(datas[i]["序号"])[-1]))
    

    print(df.shape[0])
    
    df = df.astype(str)
    Lt = list(df.groupby(key))
    n = len(Lt) // write_nt
    for i in range(write_nt):
        process = Process(target=write_df, args=(desdir, Lt[n*i:(i+1)*n]),)
        process.start()
    process = Process(target=write_df, args=(desdir, Lt[n*write_nt:len(Lt)]),)
    process.start()
